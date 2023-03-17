from django.db.models import Count
from django.http import HttpResponse

from rest_framework import generics
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from apps.dashboard.models import Report, TeamEmployee
from apps.dashboard.serializers import (
    ReportSerializer, 
    TeamEmployeeSerializer
    )

from openpyxl import Workbook
from openpyxl.utils import get_column_letter




class ReportList(generics.ListCreateAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        sent_count = Report.objects.filter(sent=True).count()
        unsent_count = Report.objects.filter(sent=False).count()
        return Response({'sent_count': sent_count, 'unsent_count': unsent_count})    



class ReportDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAdminUser]



class TeamEmployeeView(viewsets.ModelViewSet):
    queryset = TeamEmployee.objects.all()
    serializer_class = TeamEmployeeSerializer

    def get_permissions(self):
        if self.action == 'list':
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAdminUser]
        return [permission() for permission in permission_classes]




class ReportExcelView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        # Получение данных модели Report
        reports = Report.objects.all()

        # Создание новой книги Excel
        workbook = Workbook()
        sheet = workbook.active

        # Заголовки столбцов
        sheet['A1'] = 'User'
        sheet['B1'] = 'Week Start Date'
        sheet['C1'] = 'Last Week'
        sheet['D1'] = 'This Week'
        sheet['E1'] = 'Next Week'
        sheet['F1'] = 'Deadline'
        
        # Настройка ширины столбца  sheet.column_dimensions[get_column_letter(1)].width = 15
        sheet.column_dimensions[get_column_letter(2)].width = 15
        sheet.column_dimensions[get_column_letter(3)].width = 15 
        sheet.column_dimensions[get_column_letter(4)].width = 15 
        sheet.column_dimensions[get_column_letter(5)].width = 15 
        sheet.column_dimensions[get_column_letter(6)].width = 15
        # Добавление данных модели Report в таблицу
        for idx, report in enumerate(reports, start=2):
            sheet.cell(row=idx, column=1, value=report.user.username)
            sheet.cell(row=idx, column=2, value=report.week_start_date)
            sheet.cell(row=idx, column=3, value=report.last_week)
            sheet.cell(row=idx, column=4, value=report.this_week)
            sheet.cell(row=idx, column=5, value=report.next_week)
            sheet.cell(row=idx, column=6, value=report.deadline)

        # Сохранение книги в HttpResponse с нужными заголовками
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
        workbook.save(response)

        return response





